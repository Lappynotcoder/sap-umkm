"""
generate_template.py
Membuat file template Excel standar SAP-UMKM.
Jalankan sekali: python generate_template.py
Output: template_sap_umkm.xlsx (letakkan di public/template/)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def buat_template(output_path: str = "template_sap_umkm.xlsx") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data UMKM"

    # ── Warna ──
    hijau_tua  = "1A6B3A"
    hijau_muda = "E8F5EE"
    kuning     = "FFF3CD"
    abu_muda   = "F8F9FA"
    putih      = "FFFFFF"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border_tipis():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Baris 1: Judul ──
    ws.merge_cells("A1:C1")
    ws["A1"] = "TEMPLATE DATA KEUANGAN UMKM — SAP-UMKM"
    ws["A1"].font = Font(bold=True, size=13, color=putih)
    ws["A1"].fill = fill(hijau_tua)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Baris 2: Info ──
    ws.merge_cells("A2:C2")
    ws["A2"] = "Isi kolom di bawah ini. Kategori wajib: Pemasukan / HPP / Operasional"
    ws["A2"].font = Font(italic=True, color="555555", size=10)
    ws["A2"].fill = fill(hijau_muda)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Baris 3: Header Kolom ──
    headers = ["kategori", "keterangan", "nominal"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color=putih, size=11)
        cell.fill = fill(hijau_tua)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_tipis()
    ws.row_dimensions[3].height = 22

    # ── Contoh Data ──
    contoh = [
        ("Pemasukan",   "Penjualan produk utama",       8500000),
        ("Pemasukan",   "Penjualan sampingan",           1200000),
        ("HPP",         "Bahan baku utama",              2500000),
        ("HPP",         "Kemasan / packaging",            300000),
        ("Operasional", "Listrik & air",                  250000),
        ("Operasional", "Gaji karyawan",                 1500000),
        ("Operasional", "Sewa tempat",                    500000),
        ("Operasional", "Transportasi",                   150000),
        # ── Baris kosong untuk diisi user ──
        ("",            "",                                    ""),
        ("",            "",                                    ""),
        ("",            "",                                    ""),
        ("",            "",                                    ""),
        ("",            "",                                    ""),
    ]

    kategori_warna = {
        "pemasukan":   hijau_muda,
        "hpp":         "DBEAFE",
        "operasional": kuning,
    }

    for i, (kat, ket, nom) in enumerate(contoh, start=4):
        warna = kategori_warna.get(kat.lower(), putih)
        for col, val in enumerate([kat, ket, nom], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill(warna)
            c.border = border_tipis()
            c.alignment = Alignment(vertical="center")
            if col == 3 and val != "":
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")

    # ── Lebar Kolom ──
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18

    # ── Sheet Petunjuk ──
    ws2 = wb.create_sheet("Petunjuk")
    petunjuk = [
        ("PETUNJUK PENGISIAN", True),
        ("", False),
        ("1. Kolom 'kategori' wajib diisi dengan salah satu dari:", False),
        ("   - Pemasukan  → semua sumber pendapatan/penjualan", False),
        ("   - HPP        → Harga Pokok Penjualan (bahan baku, kemasan)", False),
        ("   - Operasional → biaya operasional (listrik, gaji, sewa, dll.)", False),
        ("", False),
        ("2. Kolom 'keterangan' bersifat opsional tapi disarankan diisi.", False),
        ("", False),
        ("3. Kolom 'nominal' diisi angka saja (tanpa titik/koma/Rp).", False),
        ("   Contoh: 5000000  (bukan Rp5.000.000)", False),
        ("", False),
        ("4. Boleh menambahkan baris sebanyak yang diperlukan.", False),
        ("", False),
        ("5. Simpan dalam format .xlsx atau .csv lalu upload di website.", False),
    ]
    for row_i, (text, bold) in enumerate(petunjuk, start=1):
        c = ws2.cell(row=row_i, column=1, value=text)
        c.font = Font(bold=bold, size=11 if bold else 10)
    ws2.column_dimensions["A"].width = 65

    wb.save(output_path)
    print(f"Template berhasil dibuat: {output_path}")


if __name__ == "__main__":
    buat_template()
